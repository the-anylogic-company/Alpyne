"""
Finds the optimal number of inspectors for each shift in the schedule, defined in the Excel file in the model
source, then writes them back to the file.

The strategy uses Bayesian Optimization for its sample efficiency.

Each shift has its own AnyLogicSim object and instance of the optimizer so that each iteration (round of suggestion)
can have the runs executing in parallel. To help facilitate the logic, a custom class is used to consolidate settings
and behaviors.
"""

from typing import Callable

from openpyxl import load_workbook

from alpyne.sim import AnyLogicSim
from datetime import datetime, timedelta

from bayes_opt import BayesianOptimization
from bayes_opt import UtilityFunction
from bayes_opt.util import NotUniqueError


class BCOptimizer:
    """
    Optimizer logic specific to the Border Checkpoint model
    """
    def __init__(self, round_values: bool = False, optimizer_seed: int = None):
        """
        :param round_values: Whether to round values for the run history/cache, collapsing all suggestions by the
          optimizer that round to an integer to use the exact same objective value
          (the value is rounded either way before sending to the simulation model).
        :param optimizer_seed: The seed for the optimizer to set its random state. Setting to None means a random seed.
        """
        self._round_values = round_values

        self.history = dict()
        # 'c'ar and 'b'us inspector counts hard coded based on limits defined by the sim's implementation
        self.optimizer = BayesianOptimization(f=None, pbounds={'c': (1, 6), 'b': (1, 4)}, random_state=optimizer_seed)
        # reduce alpha of the optimizer to better solve for discrete space, as per recommendation of lib's docs
        self.optimizer.set_gp_params(alpha=1e-3)
        self.utility = UtilityFunction(kind="ucb", kappa=2.5, xi=0.0)

        # keep track of how many re-attempts were made to get the next param
        #   to avoid an infinite loop.
        # reattempts are needed when the optimizer suggests a value already recorded
        #   - this namely happens when round_values is True.
        self._reattempts = 0

    def next_params_to_sim(self):
        if self._reattempts > 250:  # arbitrary limit
            raise RecursionError("Cannot find any further untried parameters to attempt")

        suggestion: dict[str, float] = self.optimizer.suggest(self.utility)

        # before running the suggestion, check to see if we have already tried it;
        # if so, recursively call this function to try to get another sample
        c, b = suggestion['c'], suggestion['b']
        key = (int(round(c)), int(round(b))) if self._round_values else (c, b)
        score = self.history.get(key)
        if score is not None:  # already recorded
            # update the optimizer with this info,
            # but don't consider it as a new iteration
            try:
                self.optimizer.register(params=suggestion, target=score)
            except NotUniqueError:  # may happen when rounding
                pass  # just ignore it
            self._reattempts += 1
            return self.next_params_to_sim()
        # found a new attempt to try; reset counter
        self._reattempts = 0

        # this is a new set of params to sim, return it to start the next iteration
        return suggestion

    def register(self, inputs: dict, score: float) -> bool:
        # register the given inputs with the underlying optimizer
        self.optimizer.register(params=inputs, target=score)

        # also store its results in the history lookup in case the inputs come up again;
        # specifically when a different floating is suggested when the actual values are set up to be rounded
        c, b = inputs['c'], inputs['b']
        key = (int(round(c)), int(round(b))) if self._round_values else (c, b)
        self.history[key] = score

        # consider this update the new best if its score matches that of the optimizer's best score
        return self.optimizer.max['target'] == score


def optimize(shifts: list[tuple[datetime, datetime]],
             objective_func: Callable[[int, int, float, float, int, int], float],
             iterations: int = 10,
             verbose: bool = False,
             **kwargs):
    """
    This will create and run parallel paired simulators and optimizers, attempting to find the best
    number of car and bus inspectors for the given shift.

    .. note:: The logic handles for cases when the optimizer runs out values to suggest which can happen
      due to the small search space, notably with the `round_values` argument being True (6*4=24 discrete options)

    :param shifts: Shifts (start and end) to optimize for
    :param objective_func: Numerical judgement for the optimizer's car & bus inputs and outputs; higher implies better
    :param iterations: How many unique parameter sets to attempt per sim/optimizer
    :param verbose: Whether to print reporting information to the console
    :param kwargs: Arguments to pass to the constructor of the optimizer
    """
    # for each shift, construct instances of the sim and an optimizer to be trained based on its results,
    #   keeping them in tuple-pairs
    sim_opt_pairs = []
    if verbose:
        print(f"CREATING OBJECTS\n{'-'*16}")
    for i, (start, stop) in enumerate(shifts):
        # lock's timeout is based on the fact that it takes ~4 (real) secs to run each sim hour, doubled for tolerance
        timeout = (stop - start).total_seconds() / 3600 * 4 * 2
        # each sim instance prevents other sims from accessing its database;
        # as this model has one, it's crucial each instance is in its own folder,
        #   which is naturally handled when passing as a zip.
        sim = AnyLogicSim("ModelExported/BorderCheckpointOptimization.zip", java_log_level=True, log_id=f"-{i + 1}", auto_lock=False,
                          engine_overrides=dict(start_date=start, stop_date=stop, seed=1),
                          lock_defaults=dict(timeout=timeout))
        opt = BCOptimizer(**kwargs)
        sim_opt_pairs.append((sim, opt))

        if verbose:
            print(f"\t#{i+1}: {start} -> {stop}")

    for iteration in range(iterations):
        if verbose:
            print(f"\nITERATION {iteration+1:02d}\n{'='*12}")

        # stores the next parameter set to try for each sim by index
        this_iter_params = dict()

        # initialize the parameters to try in this iteration,
        #   accounting for edge cases when no novel suggestions can be found
        for i, (_, opt) in enumerate(sim_opt_pairs):
            try:
                params = opt.next_params_to_sim()
            except RecursionError:
                # happens when no further untried parameters can be found;
                # just skip this one
                continue

            this_iter_params[i] = params

        # end the experiment early if all optimizers ran out of novel suggestions
        if len(this_iter_params) == 0:
            if verbose:
                print(f"\tEARLY TERMINATION")
                break

        # tell each sim to reset itself and start running (happens in the background)
        # note: because `auto_lock` was set to False, the requests here are executed nearly instantly
        for i, (sim, _) in enumerate(sim_opt_pairs):
            if i not in this_iter_params:
                # ran out of novel parameters to attempt;
                # the sim will sit idly for this round
                continue
            num_c, num_b = this_iter_params[i]['c'], this_iter_params[i]['b']
            # optimizer passes as floats; convert to rounded ints, as sim expects;
            # without this, the floats are truncated to ints, causing the upper bound values to never be attempted
            num_c, num_b = int(round(num_c)), int(round(num_b))
            sim.reset(numCarInspectors=num_c, numBusInspectors=num_b)

        # synchronize all runs by incrementally waiting til each is finished (should happen about the same time),
        #   then report the results
        for i, (sim, opt) in enumerate(sim_opt_pairs):
            if i not in this_iter_params:
                # ran out of novel parameters to attempt; skip this one
                if verbose:
                    print(f"\t#{i+1}: skipped")
                continue
            status = sim.lock()
            # get the raw (non-rounded) values
            raw_c, raw_b = this_iter_params[i]['c'], this_iter_params[i]['b']
            # convert these to their usable types
            num_c, num_b = int(round(raw_c)), int(round(raw_b))
            tis_c, tis_b = status.observation['carTISMax'], status.observation['busTISMax']
            que_c, que_b = status.observation['carsQueueing'], status.observation['busesQueueing']
            score = objective_func(num_c, num_b, tis_c, tis_b, que_c, que_b)

            new_best = opt.register(this_iter_params[i], score)

            if verbose:
                print(f"\t#{i+1}: {num_c} ({raw_c:.3f}) & {num_b} ({raw_b:.3f}) -> {tis_c:6.2f} & {tis_b:6.2f} => {score:.3f}")
                print(f"\t{status.observation['carsQueueing']} | {status.observation['busesQueueing']}")
                if new_best:
                    print("\t\tNew best!")

    # only return the found optimum for each sim
    return [opt.optimizer.max for (_, opt) in sim_opt_pairs]


if __name__ == '__main__':
    def objective(num_car_inspectors, num_bus_inspectors, car_tis, bus_tis, car_queue, bus_queue):
        # initial score definition based on penalties; i.e., smaller values == better
        score = (car_tis**2 + bus_tis**2)  # TIS has most relevance
        score *= (1 + 0.2 * ((num_car_inspectors + num_bus_inspectors)/10))  # up to 20% extra for more workers
        score *= max(1, (car_queue//50 + bus_queue//10))  # tiered penalty for excessive vehicles
        # however this optimizer is set up to maximize, so negate the score
        score *= -1
        return score

    # Read the shifts from the excel file in the source directory
    wb = load_workbook(filename=r"ModelSource\schedules.xlsx")
    ws = wb['inspectors']
    # the each shift slightly earlier to populate the model
    shifts = [(start-timedelta(hours=1), end)
              for start, end in ws.iter_rows(min_row=2, max_col=2, values_only=True)]

    shift_bests = optimize(
        shifts,
        objective,
        iterations=10,
        verbose=True,
        round_values=True,  # uses rounded nums for historical cache, ensuring each iteration has unique attempts
        optimizer_seed=1  # fixed value will use the initial guess for all but will deviate based on results
    )

    # print out and store the bests for each in lists
    print("\nRESULTS\n=======")
    for i, best in enumerate(shift_bests):
        car_value, bus_value = round(best['params']['c']), round(best['params']['b'])

        # car value in column C, bus in D;
        # +2 b/c excel starts at row 1, where is where headers are
        ws[f'C{i+2}'].value = car_value
        ws[f'D{i+2}'].value = bus_value

        print(f"SHIFT {i+1} BEST: c={car_value:.0f}, b={bus_value:.0f} | {best}")

    # save changes to the original excel file
    wb.save(r"ModelSource\schedules.xlsx")

    # now, create another instance just to run a full 24 hours to test the values
    # (0 inspectors is interpreted as using the schedule)
    sim = AnyLogicSim("ModelExported/BorderCheckpointOptimization.zip",
                      engine_overrides=dict(start_date=datetime(2024, 1, 1),
                                            stop_date=datetime(2024, 1, 2),
                                            seed=1),
                      lock_defaults=dict(timeout=100))
    status = sim.reset(numCarInspectors=0, numBusInspectors=0)
    print(status.observation)

    # you can also go re-run the animated model in AnyLogic and see the optimized values!




